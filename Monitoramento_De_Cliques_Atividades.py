import tkinter as tk  # Importa a biblioteca Tkinter para criar interfaces gráficas
import openpyxl  # Biblioteca para manipulação de arquivos Excel
import pandas as pd  # Biblioteca para manipulação de dados
import os  # Biblioteca para manipulação do sistema operacional
import time  # Biblioteca para manipulação de tempo
import threading  # Biblioteca para manipulação de threads
import logging  # Biblioteca para geração de logs
import win32api, win32con  # Módulos para interação com a API do Windows
import win32ts  # Para monitorar eventos de sessão no Windows
from tkinter import ttk, messagebox  # Importa componentes adicionais do Tkinter
from pynput.mouse import Listener  # Biblioteca para capturar eventos do mouse
from openpyxl import load_workbook, Workbook  # Métodos para manipulação de planilhas Excel
from pywinauto import Application, Desktop  # Biblioteca para automação de interfaces do Windows
from openpyxl.styles import NamedStyle, Font  # Estilos para formatação de planilhas Excel
from openpyxl.worksheet.table import Table, TableStyleInfo  # Criação e estilo de tabelas em Excel

# Configuração inicial do aplicativo Tkinter
app = tk.Tk()
app.title("Monitor de Cliques")  # Define o título da janela
app.geometry("600x300")  # Define o tamanho da janela

# Variáveis globais
evento_parada = threading.Event()  # Evento para controle de threads
listener_ativo = False  # Variável que indica se a escuta do mouse está ativa
listener = None  # Objeto que armazenará o Listener
nome_arquivo = ""  # Nome do arquivo Excel

# DataFrame para armazenar os dados
df = pd.DataFrame()
linha_atual = 2  # Linha atual do Excel para inserção de novos dados
ws = None  # Variável global para a aba da planilha
wb = None  # Variável global para o arquivo Excel

# Lista de atividades disponíveis para seleção pelo usuário
listAtividades = [
    "1 - Atendimento De Um Chamado Novo",
    "2 - Inclusão Do Pedido De Coleta",
    "3 - Atualização De Um Chamado",
    "4 - Abertura De Uma OS",
    "5 - Solicitação Via TEAMS",
    "6 - Contato Com Repetec",
    "7 - Pedido De Envio De Peça",
    "8 - Pedido De Cabo De Dados",
    "9 - Pedido De RMA",
    "10 - Pedido De Prioridade Junto Ao LAB",
    "11 - Pedido De Prioridade Junto A LOG",
    "12 - Contato Com A Loja",
    "13 - Contato Com O Motorista",
    "14 - Contato Com O CD/ Loja",
    "15 - Atendimento De Uma RITM",
    "16 - Encerramento Do Chamado",
    "17 - Pedido De Retorno Peça/ Repetec",
    "18 - E-mail Para A Zebra",
    "19 - Contato Com A Loja Para Confirmação De Recebimento",
    "20 - Gerar Orçamento",
    "21 - Atualizar OS",
    "22 - Enviar Orçamento",
    "23 - Cobrar Aprovação",
    "24 - Gerar Relatório De Faturamento",
    "25 - Conferência Relatório Garantia",
    "26 - Cobrar Compras",
    "27 - Conferir Pendências Recebimento",
    "28 - Contato Com Recebimento",
    "29 - Contato Com Lab Para Pendência",
    "30 - Contato Com Prissing",
    "31 - Contato Com Comercial",
    "32 - Contato Com Estoque" 
]

# Configuração da interface gráfica
lb_atividades = tk.Label(app, text="Selecione a atividade para contagem!")  # Rótulo de instrução
lb_atividades.pack(pady=10)  # Adiciona espaço ao redor do rótulo

# ComboBox para selecionar atividades
cd_atividades = ttk.Combobox(app, values=listAtividades, width=50)
cd_atividades.set("Selecione uma atividade")  # Define valor inicial
cd_atividades.pack(pady=10)  # Adiciona o ComboBox à interface


# Função para verificar se a planilha está aberta antes de salvar
def verificar_planilha_aberta(nome_arquivo):
    if os.path.exists(nome_arquivo):  # Verifica se o arquivo existe
        try:
            with open(nome_arquivo, "r+"):  # Tenta abrir para leitura e escrita
                return False  # Arquivo não está em uso
        except PermissionError:
            return True  # Arquivo está em uso
    return False  # Arquivo não existe

# Variável para armazenar a última atividade selecionada
atividade_anterior = "Selecione uma atividade"

# Função para bloquear a troca de atividade durante a contagem
def bloquear_troca_atividade(event):
    global atividade_anterior
    if listener_ativo:
        messagebox.showwarning(
            "Ação não permitida",
            "Ação não permitida! Necessário finalizar a contagem da primeira atividade para iniciar outra."
        )
        cd_atividades.set(atividade_anterior)  # Restaura a última atividade
    else:
        atividade_anterior = cd_atividades.get()  # Atualiza a atividade atual

# Associa o evento de mudança de seleção no Combobox
cd_atividades.bind("<<ComboboxSelected>>", bloquear_troca_atividade)

# Configuração de mensagens no app
mensagem_label = tk.Label(app, font=("Arial", 12), pady=10)

# Configuração de log para depuração
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Função para criar ou carregar o DataFrame e planilha
def criar_ou_abrir_planilha():
    global nome_arquivo, df, ws, wb
    data_hoje = time.strftime('%d%m%y')
    nome_arquivo = f"Rel_Monitoramento_Cliques{data_hoje}.xlsx"

    if os.path.exists(nome_arquivo):
        # Carregar o arquivo Excel existente
        wb = load_workbook(nome_arquivo)

        # Verificar se a aba "Contagem" existe
        if "Monitoramento" in wb.sheetnames:
            ws = wb["Monitoramento"]
        else:
            ws = wb.create_sheet("Monitoramento")
            ws.append(["X", "Y", "Data", "Hora", "Contagem Segundos", "Tipo de Navegação", "Nome da Janela", "Nome do Clique", "Atividade"])

        # Carregar os dados da planilha para o DataFrame
        try:
            df = pd.read_excel(nome_arquivo, sheet_name="Monitoramento")

            # **Garantir que os dados antigos são carregados corretamente na planilha**
            for _, row in df.iterrows():
                ws.append(row.tolist())

        except ValueError:
            print("A aba 'Contagem' não possui dados ou está vazia.")
            df = pd.DataFrame(columns=["X", "Y", "Data", "Hora", "Contagem Segundos", "Tipo de Navegação", "Nome da Janela", "Nome do Clique", "Atividade"])

    else:
        # Criar um novo arquivo Excel e aba
        wb = Workbook()
        ws = wb.active
        ws.title = "Monitoramento"

        colunas = ["X", "Y", "Data", "Hora", "Contagem Segundos", "Tipo de Navegação", "Nome da Janela", "Nome do Clique", "Atividade"]
        df = pd.DataFrame(columns=colunas)
        ws.append(colunas)  # Adicionar cabeçalho na nova planilha

        # Salvar o arquivo para evitar erros posteriores
        wb.save(nome_arquivo)

    return df

#Inicializar o DataFrame
df = criar_ou_abrir_planilha()
# Funções auxiliares
def obter_janela_ativa():
    try:
        janela_ativa = Application(backend='uia').connect(active_only=True).top_window()
        return janela_ativa.window_text() if janela_ativa else "Janela Desconhecida"
    except Exception as e:
        logging.warning(f"Erro ao capturar janela ativa: {e}")
        return "Erro na Captura da Janela"

def obter_nome_controle(x, y):
    try:
        elemento = Desktop(backend='uia').from_point(x, y)
        return elemento.window_text() or "Controle sem texto" if elemento else "Controle Desconhecido"
    except Exception as e:
        logging.warning(f"Erro ao capturar o controle: {e}")
        return "Erro ao Capturar Controle"

# Função chamada ao clicar
def ao_clicar(x, y, botao, pressionado):
    global df, linha_atual

    if pressionado:
        data = time.strftime('%Y-%m-%d')
        hora_completa = time.strftime('%H:%M:%S')

        # Capturar informações da janela ativa
        titulo_janela = obter_janela_ativa()
        tipo_janela = "Aplicativo" if " - " not in titulo_janela.lower() else "Navegador"
        nome_controle = obter_nome_controle(x, y)

        logging.info(f'Clique registrado em ({x}, {y}) na janela: "{titulo_janela}", Controle: "{nome_controle}"')

        # Determinar a atividade registrada
        atividade = cd_atividades.get()
        if titulo_janela == "Tela de Bloqueio padrão do Windows":
            atividade = "Tela de Bloqueio padrão do Windows"

        # Calcular diferença de tempo
        if not df.empty:
            ultima_hora = df.iloc[-1]["Hora"]
            ultima_hora = pd.to_datetime(ultima_hora, format="%H:%M:%S").time()
            hora_atual = pd.to_datetime(hora_completa, format="%H:%M:%S").time()
            diferenca = (pd.to_datetime(hora_atual.strftime('%H:%M:%S')) - pd.to_datetime(ultima_hora.strftime('%H:%M:%S'))).seconds
            contagem_segundos = f"{diferenca // 3600:02}:{(diferenca % 3600) // 60:02}:{diferenca % 60:02}"
        else:
            contagem_segundos = "00:00:00"

        # Adicionar nova linha ao DataFrame
        novo_dado = pd.DataFrame({
            "X": [x], "Y": [y], "Data": [data], "Hora": [hora_completa],
            "Contagem Segundos": [contagem_segundos], "Tipo de Navegação": [tipo_janela],
            "Nome da Janela": [titulo_janela], "Nome do Clique": [nome_controle], "Atividade": [atividade]
        })

        df = pd.concat([df, novo_dado], ignore_index=True)
        df["Hora"] = pd.to_datetime(df["Hora"], format="%H:%M:%S").dt.time

        # Registrar na planilha
        try:
            ws.append([x, y, data, hora_completa, contagem_segundos, tipo_janela, titulo_janela, nome_controle, atividade])
            wb.save(nome_arquivo)
            linha_atual += 1
        except Exception as e:
            logging.error(f"Erro ao registrar na planilha: {e}")

        wb.save(nome_arquivo)
        wb.close()


# Função para salvar a planilha
def salvar_planilha():
    if verificar_planilha_aberta(nome_arquivo):
        messagebox.showerror("Erro", "O arquivo já está aberto! Feche-o antes de salvar.")
        return
    
    df.to_excel(nome_arquivo, index=False)
    logging.info("Planilha salva com sucesso.")
   


def carregar_dados(nome_arquivo):

    if os.path.exists(nome_arquivo):
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb.active        
        data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
        colunas = ["X", "Y", "Data", "Hora", "Contagem Segundos", "Tipo de Navegação", "Nome da Janela", "Nome do Clique", "Atividade"]

        ws.title = "Monitoramento"
        # Ajustar larguras das colunas
        ws.column_dimensions["A"].width = 6.00
        ws.column_dimensions["B"].width = 6.00
        ws.column_dimensions["C"].width = 11.00
        ws.column_dimensions["D"].width = 9.00
        ws.column_dimensions["E"].width = 18.14
        ws.column_dimensions["F"].width = 17.00
        ws.column_dimensions["G"].width = 74.00
        ws.column_dimensions["H"].width = 30.71
        ws.column_dimensions["I"].width = 48.00

        # Definir estilo para o título (cabeçalho)
        titulo_fonte = Font(color="FFFFFF", bold=True)  # Fonte branca e negrito
        
        # Aplicar o estilo de título para a primeira linha (somente os cabeçalhos)
        for cell in ws[1]:  # Iterar sobre todas as células da primeira linha (títulos)
            cell.font = titulo_fonte

        # Definir intervalo dinâmico para a tabela (sem limite de linhas)
        tabela = Table(displayName="TabelaContagem", ref=f"A1:I29000")  # Usar a última linha da planilha
        estilo = TableStyleInfo(
            name="TableStyleMedium5",  # Estilo de tabela média 5
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)

        # Salvar o arquivo Excel com a estrutura inicial
        wb.save(nome_arquivo)

        return pd.DataFrame(data, columns=colunas), wb
    else:
        return df, None

def somar_tempo_por_atividade(df):
    df["Contagem Segundos"] = pd.to_timedelta(df["Contagem Segundos"])
    return df.groupby("Atividade")["Contagem Segundos"].sum().reset_index()

def salvar_relatorio(wb, relatorio):
    if verificar_planilha_aberta(nome_arquivo):
        messagebox.showerror("Erro", "O arquivo já está aberto! Feche-o antes de salvar.")
        return

    if "Relatório" in wb.sheetnames:
        del wb["Relatório"]
    
    nova_aba = wb.create_sheet("Relatório")
    nova_aba.append(["Atividade", "Tempo Gasto (hh:mm:ss)"])

    # Ajustar colunas
    nova_aba.column_dimensions["A"].width = 40.86
    nova_aba.column_dimensions["B"].width = 33.43

    estilo_hora = NamedStyle(name="estilo_hora", number_format="[h]:mm:ss")
    if "estilo_hora" not in wb.named_styles:
        wb.add_named_style(estilo_hora)

    tabela = Table(displayName="TabelaRelatório", ref="A1:B32")
    estilo = TableStyleInfo(
        name="TableStyleMedium5",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    tabela.tableStyleInfo = estilo
    nova_aba.add_table(tabela)

    for _, row in relatorio.iterrows():
        nova_aba.append([row["Atividade"], row["Contagem Segundos"]])
        nova_aba[f"B{nova_aba.max_row}"].style = estilo_hora

    wb.save(nome_arquivo)
    messagebox.showinfo("Sucesso", "Relatório gerado com sucesso!")


def gerar_relatorio():
    df, wb = carregar_dados(nome_arquivo)
    if wb:
        relatorio = somar_tempo_por_atividade(df)
        salvar_relatorio(wb, relatorio)
# Função para iniciar a contagem de cliques
def iniciar_contagem():
    global listener_ativo, listener

    if verificar_planilha_aberta(nome_arquivo):
        messagebox.showerror("Erro", "O arquivo já está aberto! Feche-o antes de começar a contagem.")
        return
    # Verificar se já há uma contagem em andamento
    if verificar_contagem_em_andamento():
        return  # Interromper a função se já houver uma contagem ativa

    # Verificar se o usuário selecionou uma atividade válida
    if cd_atividades.get() == "Selecione uma atividade":
        messagebox.showwarning("Aviso", "Você não selecionou a atividade!")
        return

    # Iniciar o listener de cliques
    if not listener_ativo:
        listener = Listener(on_click=ao_clicar)
        listener.start()
        listener_ativo = True
        logging.info("Contagem de cliques iniciada.")
    
    mensagem_label.config(text="Clique em qualquer lugar para iniciar a contagem!")
    mensagem_label.pack()
    app.bind("<Button-1>", ocultar_mensagem)


# Função para monitorar bloqueio de tela
def monitorar_bloqueio_tela():
    global df, linha_atual
    while not evento_parada.is_set():
        try:
            session_id = win32api.WTSGetActiveConsoleSessionId()
            estado_sessao = win32ts.WTSQuerySessionInformation(
                win32ts.WTS_CURRENT_SERVER_HANDLE,
                session_id,
                win32ts.WTSConnectState
            )

            # Se a sessão estiver desconectada, a tela está bloqueada
            if estado_sessao == win32ts.WTSDisconnected:
                logging.info("Tela bloqueada detectada.")
                
                # Registrar evento de bloqueio na planilha
                registrar_bloqueio_tela()

                # Finalizar contagem e exibir mensagem
                finalizar_contagem()
                app.after(0, lambda: messagebox.showinfo(
                    "Bloqueio de Tela",
                    "Relatório finalizado devido ao bloqueio de tela. Selecione a atividade e clique em iniciar contagem para iniciar novamente."
                ))
                break

        except Exception as e:
            logging.error(f"Erro ao monitorar bloqueio de tela: {e}")

        time.sleep(1)
def registrar_bloqueio_tela():
    global df, linha_atual
    data = time.strftime('%Y-%m-%d')
    hora_completa = time.strftime('%H:%M:%S')

    # Substituir atividade para "Bloqueio de Tela"
    atividade_bloqueio = "Bloqueio de Tela"

    novo_dado = pd.DataFrame({
        "X": [0], "Y": [0], "Data": [data], "Hora": [hora_completa],
        "Contagem Segundos": ["00:00:00"], "Tipo de Navegação": ["Sistema"],
        "Nome da Janela": ["Tela Bloqueada"], "Nome do Clique": ["N/A"], "Atividade": [atividade_bloqueio]
    })

    df = pd.concat([df, novo_dado], ignore_index=True)

    # Adicionar à planilha
    ws.append([0, 0, data, hora_completa, "00:00:00", "Sistema", "Tela Bloqueada", "N/A", atividade_bloqueio])
    wb.save(nome_arquivo)
    logging.info("Registro de bloqueio de tela adicionado com sucesso.")


def verificar_contagem_em_andamento():
    if listener_ativo:
        messagebox.showwarning(
            "Contagem em andamento",
            "Você já iniciou a contagem de uma atividade. Favor finalizar a 1º contagem para iniciar a próxima."
        )
        return True
    return False

# Função para ocultar mensagens de status
def ocultar_mensagem(event):
    mensagem_label.pack_forget()
    app.unbind("<Button-1>")

# Função para finalizar a contagem de cliques
def finalizar_contagem():
    global listener_ativo, listener

    if not listener_ativo:
        messagebox.showwarning(
            "Aviso",
            "Você ainda não selecionou a atividade e nem clicou em iniciar a contagem. "
            "Este botão é apenas para finalizar o relatório."
        )
        return

    if verificar_planilha_aberta(nome_arquivo):
        messagebox.showerror("Erro", "O arquivo já está aberto! Feche-o antes de salvar.")
        return

    # Parar a contagem de cliques
    listener.stop()
    listener_ativo = False

    # Salvar os dados e gerar o relatório
    salvar_planilha()
    gerar_relatorio()
    logging.info("Contagem de cliques finalizada.")

    # Atualizar a interface com a mensagem de sucesso
    mensagem_label.pack_forget()  # Esconde mensagens anteriores
    mensagem_label.config(text="Relatório criado com sucesso!")
    mensagem_label.pack()
    
    # Ocultar mensagem ao clicar na janela
    app.bind("<Button-1>", ocultar_mensagem)

# Função para exibir informações sobre o aplicativo
def exibir_info():
    info_texto = "Monitor de Cliques - Versão 1.0\nDesenvolvido por Pamella Barros e Robson Calheira"
    info_janela = tk.Toplevel(app)
    info_janela.title("Sobre")
    tk.Label(info_janela, text=info_texto, font=("Arial", 12), padx=10, pady=10).pack()
    tk.Button(info_janela, text="Fechar", command=info_janela.destroy).pack(pady=10)
    

# Adicionar botões de controle
btn_iniciar = tk.Button(app, text="Iniciar Contagem", command=iniciar_contagem)
btn_iniciar.pack(pady=5)

btn_finalizar = tk.Button(app, text="Finalizar Contagem", command=finalizar_contagem)
btn_finalizar.pack(pady=5)

btn_info = tk.Button(app, text="Sobre", command=exibir_info)
btn_info.pack(pady=10)

# Configuração para fechar o aplicativo
# Função para fechar o aplicativo com mensagem de aviso
def ao_fechar():
    if listener_ativo:
        # Exibir mensagem de aviso se a contagem ainda estiver ativa
        messagebox.showwarning(
            "Aviso",
            "Você ainda não finalizou seu relatório. Finalize a contagem para poder fechar o programa!"
        )
    else:
        # Encerrar o programa se não houver contagem ativa
        evento_parada.set()
        app.destroy()


app.protocol("WM_DELETE_WINDOW", ao_fechar)

# Executar a interface
app.mainloop()